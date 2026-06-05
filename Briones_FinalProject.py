import tkinter as tk
from tkinter import messagebox, ttk
from datetime import datetime
import openpyxl as op
import os

# ==========================================
# DATABASE BACKEND LOGIC
# ==========================================

def create_excel_db():
    """Create Excel file with reservation headers if it doesn't exist"""
    if not os.path.exists("briones_Database.xlsx"):
        wb = op.Workbook()
        ws = wb.active
        ws.title = "ReservationRecords"
        ws.append([
            "Reservation ID", "Customer Name", "Facility Name", "Phone", "Email", 
            "Membership Type", "Duration (Hours)", "Hourly Rate", "Total Amount", "Status", "Reservation Date"
        ])
        
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
            
        wb.save("briones_Database.xlsx")

def display_excel():
    """Load records from Excel and display them in the Treeview table"""
    try:
        workbook = op.load_workbook("briones_Database.xlsx")
        sheet = workbook.active
        
        for content in table.get_children():
            table.delete(content)
            
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                table.insert("", tk.END, values=row)
    except Exception as e:
        messagebox.showerror("Error", f"Could not load data: {e}")

def generate_id():
    """Generate automatic ID like RES001, RES002..."""
    if not os.path.exists("briones_Database.xlsx"):
        return "RES001"
        
    workbook = op.load_workbook("briones_Database.xlsx")
    sheet = workbook.active
    last_row = sheet.max_row
    
    if last_row == 1:
        return "RES001"
        
    last_id = sheet.cell(row=last_row, column=1).value
    if last_id and str(last_id).startswith("RES"):
        try:
            num = int(last_id[3:]) + 1
            return f"RES{num:03d}"
        except ValueError:
            return "RES001"
    return "RES001"

# ==========================================
# COMPUTATION & VALIDATION
# ==========================================

def calculate_total(event=None):
    """AUTOMATIC COMPUTATION: Calculates Total Amount = Duration * Hourly Rate"""
    duration = duration_entry.get().strip()
    rate = rate_entry.get().strip()
    
    if duration.isdigit() and rate.replace('.', '', 1).isdigit():
        total = float(duration) * float(rate)
        total_entry.config(state='normal')
        total_entry.delete(0, tk.END)
        total_entry.insert(0, f"{total:.2f}")
        total_entry.config(state='readonly')
    else:
        total_entry.config(state='normal')
        total_entry.delete(0, tk.END)
        total_entry.config(state='readonly')

def validation():
    """Validates form inputs before saving or updating"""
    customer = customer_entry.get().strip()
    facility = facility_entry.get().strip()
    phone = phone_entry.get().strip()
    email = email_entry.get().strip()
    membership = membership_combo.get()
    duration = duration_entry.get().strip()
    rate = rate_entry.get().strip()
    status = status_combo.get()
    
    if not (customer and facility and phone and email and membership and duration and rate and status):
        messagebox.showerror("Error", "All fields are required!")
        return False
        
    if not duration.isdigit():
        messagebox.showerror("Error", "Duration must be a whole number!")
        return False
        
    if not rate.replace('.', '', 1).isdigit():
        messagebox.showerror("Error", "Hourly Rate must be a valid number!")
        return False
        
    return True

# ==========================================
# CRUD OPERATIONS (Create, Read, Update, Delete)
# ==========================================

def select_record(event=None):
    """Populate entry boxes when a row is clicked in the table"""
    selected = table.focus()
    if not selected:
        return
        
    values = table.item(selected, "values")
    if not values:
        return
        
    # Clear all entries first
    clear_fields()
    
    res_id_entry.config(state='normal')
    res_id_entry.delete(0, tk.END)
    customer_entry.delete(0, tk.END)
    facility_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)
    duration_entry.delete(0, tk.END)
    rate_entry.delete(0, tk.END)
    total_entry.config(state='normal')
    total_entry.delete(0, tk.END)
    date_entry.delete(0, tk.END)
    
    # Fill entries with selected values
    res_id_entry.insert(0, values[0])
    res_id_entry.config(state='readonly')
    customer_entry.insert(0, values[1])
    facility_entry.insert(0, values[2])
    phone_entry.insert(0, values[3])
    email_entry.insert(0, values[4])
    membership_combo.set(values[5])
    duration_entry.insert(0, values[6])
    rate_entry.insert(0, values[7])
    total_entry.insert(0, values[8])
    total_entry.config(state='readonly')
    status_combo.set(values[9])
    date_entry.insert(0, values[10])

def append_excel():
    """Add a new reservation record"""
    if not validation():
        return
        
    customer = customer_entry.get().strip().title()
    facility = facility_entry.get().strip().title()
    phone = phone_entry.get().strip()
    email = email_entry.get().strip().lower()
    membership = membership_combo.get()
    duration = int(duration_entry.get().strip())
    rate = float(rate_entry.get().strip())
    total = duration * rate
    status = status_combo.get()
    date_val = date_entry.get().strip()
    
    if not date_val:
        date_val = datetime.now().strftime("%Y-%m-%d %H:%M")
        
    new_id = generate_id()
    
    workbook = op.load_workbook("briones_Database.xlsx")
    sheet = workbook.active
    
    sheet.append([new_id, customer, facility, phone, email, membership, duration, rate, total, status, date_val])
    workbook.save("briones_Database.xlsx")
    
    messagebox.showinfo("Success", f"Reservation added successfully!\nID: {new_id}")
    clear_fields()
    display_excel()

def update_data():
    """Update selected reservation data"""
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error", "Please select a record to update.")
        return
        
    values = table.item(selected, "values")
    record_id = values[0]
    
    if not validation():
        return
        
    confirm = messagebox.askyesno("Confirm Update", "Are you sure you want to update this reservation?")
    if not confirm:
        return
        
    customer = customer_entry.get().strip().title()
    facility = facility_entry.get().strip().title()
    phone = phone_entry.get().strip()
    email = email_entry.get().strip().lower()
    membership = membership_combo.get()
    duration = int(duration_entry.get().strip())
    rate = float(rate_entry.get().strip())
    total = duration * rate
    status = status_combo.get()
    date_val = date_entry.get().strip()
    
    if not date_val:
        date_val = datetime.now().strftime("%Y-%m-%d %H:%M")
        
    workbook = op.load_workbook("briones_Database.xlsx")
    sheet = workbook.active
    
    found = False
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == record_id:
            row[1].value = customer
            row[2].value = facility
            row[3].value = phone
            row[4].value = email
            row[5].value = membership
            row[6].value = duration
            row[7].value = rate
            row[8].value = total
            row[9].value = status
            row[10].value = date_val
            found = True
            break
            
    if found:
        workbook.save("briones_Database.xlsx")
        messagebox.showinfo("Success", "Reservation updated successfully!")
        clear_fields()
        display_excel()
    else:
        messagebox.showerror("Error", "Record not found.")

def delete_data():
    """Delete selected reservation record"""
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error", "Please select a record to delete.")
        return
        
    values = table.item(selected, "values")
    record_id = values[0]
    
    confirm = messagebox.askyesno("Confirm Delete", f"Delete Reservation ID {record_id}? This cannot be undone!")
    if not confirm:
        return
        
    workbook = op.load_workbook("briones_Database.xlsx")
    sheet = workbook.active
    
    deleted = False
    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row[0].value == record_id:
            sheet.delete_rows(i, 1)
            deleted = True
            break
            
    if deleted:
        workbook.save("briones_Database.xlsx")
        messagebox.showinfo("Success", "Reservation deleted successfully!")
        clear_fields()
        display_excel()
    else:
        messagebox.showerror("Error", "Record not found.")

def clear_fields():
    """Clear all form input boxes"""
    res_id_entry.config(state='normal')
    res_id_entry.delete(0, tk.END)
    res_id_entry.config(state='readonly')
    
    customer_entry.delete(0, tk.END)
    facility_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)
    membership_combo.set('')
    duration_entry.delete(0, tk.END)
    rate_entry.delete(0, tk.END)
    
    total_entry.config(state='normal')
    total_entry.delete(0, tk.END)
    total_entry.config(state='readonly')
    
    status_combo.set('')
    date_entry.delete(0, tk.END)
    date_entry.insert(0, datetime.now().strftime("%Y-%m-%d %H:%M"))

# ==========================================
# UI INTERFACE LAYOUT (Tkinter Window)
# ==========================================

window = tk.Tk()
window.title("Sports Facility Reservation System")
window.configure(bg="lightblue")
window.geometry("1150x650")

create_excel_db()

# Main Title Label
title = tk.Label(
    window, text="Sports Facility Reservation System", 
    font=("Times New Roman", 18, "bold"), bg="lightblue"
)
title.grid(row=0, column=0, columnspan=8, pady=(10, 5))

# Input Frame Wrapper
genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=8, padx=10, pady=10, sticky="ew")

# ROW 1 INSIDE GENFRAME
res_id_label = tk.Label(genframe, text="Reservation ID:", font=("Poppins", 10, "italic"), bg="lightblue")
res_id_label.grid(row=0, column=0, padx=(10, 0), pady=(10, 0), sticky="w")
res_id_entry = tk.Entry(genframe, font=("Poppins", 12), width=12, state='readonly')
res_id_entry.grid(row=0, column=1, padx=(5, 10), pady=(10, 0))

customer_label = tk.Label(genframe, text="Customer Name:", font=("Poppins", 10, "italic"), bg="lightblue")
customer_label.grid(row=0, column=2, padx=(10, 0), pady=(10, 0), sticky="w")
customer_entry = tk.Entry(genframe, font=("Poppins", 12), width=22)
customer_entry.grid(row=0, column=3, columnspan=2, padx=(5, 10), pady=(10, 0))

facility_label = tk.Label(genframe, text="Facility Name:", font=("Poppins", 10, "italic"), bg="lightblue")
facility_label.grid(row=0, column=5, padx=(10, 0), pady=(10, 0), sticky="w")
facility_entry = tk.Entry(genframe, font=("Poppins", 12), width=18)
facility_entry.grid(row=0, column=6, columnspan=2, padx=(5, 10), pady=(10, 0))

# ROW 2 INSIDE GENFRAME
phone_label = tk.Label(genframe, text="Phone:", font=("Poppins", 10, "italic"), bg="lightblue")
phone_label.grid(row=1, column=0, padx=(10, 0), pady=(10, 0), sticky="w")
phone_entry = tk.Entry(genframe, font=("Poppins", 12), width=12)
phone_entry.grid(row=1, column=1, padx=(5, 10), pady=(10, 0))

email_label = tk.Label(genframe, text="Email:", font=("Poppins", 10, "italic"), bg="lightblue")
email_label.grid(row=1, column=2, padx=(10, 0), pady=(10, 0), sticky="w")
email_entry = tk.Entry(genframe, font=("Poppins", 12), width=22)
email_entry.grid(row=1, column=3, columnspan=2, padx=(5, 10), pady=(10, 0))

membership_label = tk.Label(genframe, text="Membership:", font=("Poppins", 10, "italic"), bg="lightblue")
membership_label.grid(row=1, column=5, padx=(10, 0), pady=(10, 0), sticky="w")
membership_combo = ttk.Combobox(
    genframe, font=("Poppins", 12), width=15, state="readonly",
    values=["Regular", "VIP", "Premium", "Guest"]
)
membership_combo.grid(row=1, column=6, padx=(5, 10), pady=(10, 0))

# ROW 3 INSIDE GENFRAME
duration_label = tk.Label(genframe, text="Duration (Hours):", font=("Poppins", 10, "italic"), bg="lightblue")
duration_label.grid(row=2, column=0, padx=(10, 0), pady=(10, 0), sticky="w")
duration_entry = tk.Entry(genframe, font=("Poppins", 12), width=12)
duration_entry.grid(row=2, column=1, padx=(5, 10), pady=(10, 0))
duration_entry.bind("<KeyRelease>", calculate_total)

rate_label = tk.Label(genframe, text="Hourly Rate ($):", font=("Poppins", 10, "italic"), bg="lightblue")
rate_label.grid(row=2, column=2, padx=(10, 0), pady=(10, 0), sticky="w")
rate_entry = tk.Entry(genframe, font=("Poppins", 12), width=12)
rate_entry.grid(row=2, column=3, padx=(5, 10), pady=(10, 0))
rate_entry.bind("<KeyRelease>", calculate_total)

total_label = tk.Label(genframe, text="Total Amount:", font=("Poppins", 10, ["bold", "italic"]), bg="lightblue")
total_label.grid(row=2, column=4, padx=(10, 0), pady=(10, 0), sticky="w")
total_entry = tk.Entry(genframe, font=("Poppins", 12), width=12, state='readonly', bg="white")
total_entry.grid(row=2, column=5, padx=(5, 10), pady=(10, 0))

status_label = tk.Label(genframe, text="Status:", font=("Poppins", 10, "italic"), bg="lightblue")
status_label.grid(row=2, column=6, padx=(10, 0), pady=(10, 0), sticky="w")
status_combo = ttk.Combobox(genframe, font=("Poppins", 12), width=15, state="readonly", values=["Active", "Inactive"])
status_combo.grid(row=2, column=7, padx=(5, 10), pady=(10, 0))

# ROW 4 INSIDE GENFRAME (Date Picker Box)
date_label = tk.Label(genframe, text="Date Added:", font=("Poppins", 10, "italic"), bg="lightblue")
date_label.grid(row=3, column=0, padx=(10, 0), pady=(10, 10), sticky="w")
date_entry = tk.Entry(genframe, font=("Poppins", 12), width=18)
date_entry.grid(row=3, column=1, columnspan=2, padx=(5, 10), pady=(10, 10), sticky="w")
date_entry.insert(0, datetime.now().strftime("%Y-%m-%d %H:%M"))

# ==========================================
# BUTTON CONTROL BAR
# ==========================================

btn_frame = tk.Frame(window, bg="lightblue")
btn_frame.grid(row=2, column=0, columnspan=8, pady=(5, 10))

submit_button = tk.Button(btn_frame, text="Add Reservation", font=("Poppins", 12, "bold"), bg="lightpink", command=append_excel, width=16)
submit_button.grid(row=0, column=0, padx=5)

update_btn = tk.Button(btn_frame, text="Update", font=("Poppins", 12, "bold"), bg="yellow", command=update_data, width=14)
update_btn.grid(row=0, column=1, padx=5)

delete_btn = tk.Button(btn_frame, text="Delete", font=("Poppins", 12, "bold"), bg="red", fg="white", command=delete_data, width=14)
delete_btn.grid(row=0, column=2, padx=5)

clear_btn = tk.Button(btn_frame, text="Clear", font=("Poppins", 12, "bold"), bg="gray", fg="white", command=clear_fields, width=14)
clear_btn.grid(row=0, column=3, padx=5)

refresh_btn = tk.Button(btn_frame, text="Refresh", font=("Poppins", 12, "bold"), bg="cyan", command=display_excel, width=14)
refresh_btn.grid(row=0, column=4, padx=5)

# ==========================================
# THE TREEVIEW TABLE DATA VIEW
# ==========================================

columns = (
    "ID", "Customer Name", "Facility Name", "Phone", "Email", 
    "Membership", "Duration", "Hourly Rate", "Total Amount", "Status", "Date"
)
table = ttk.Treeview(window, columns=columns, show="headings")

# Setup headers and alignments
for col in columns:
    table.heading(col, text=col)
    if col in ["ID", "Duration", "Hourly Rate", "Total Amount", "Status"]:
        table.column(col, width=90, anchor=tk.CENTER)
    else:
        table.column(col, width=110, anchor=tk.W)

table.grid(row=3, column=0, columnspan=8, padx=10, pady=(0, 10), sticky="nsew")
table.bind("<<TreeviewSelect>>", select_record)

# Window Responsiveness configs
window.grid_rowconfigure(3, weight=1)
window.grid_columnconfigure(0, weight=1)

# App Boot Init initialization
display_excel()
window.mainloop()
